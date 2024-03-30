# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import os
import pandas as pd
from openpyxl import Workbook
import yfinance as yf
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import minimize
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Define the list of ticker symbols
ticker_symbols = ["INTC", "GOOG", "AAPL", "MSFT", "KO", "TSLA", "NVDA", "AMZN", "AVGO", "V", "WMT", "UNH", "HD", "OXY", "JNJ", "COST", "CVX", "AMD", "MCD", "CSCO", "ABT", "CAT", "TXN", "PFE", "PM", "UPS", "T", "VZ", "DE", "BA", "LMT", "SBUX"]

# Initialize an empty DataFrame to store the DuPont Analysis breakdown
dupont_breakdown_df = pd.DataFrame(columns=["Net_Income_to_EBT", "EBT_to_EBIT", "EBIT_to_Revenues", "Revenues_to_Assets", "Assets_to_Equity"], index=ticker_symbols)

# Iterate over each ticker symbol
for ticker_symbol in ticker_symbols:
    try:
        # Fetch data for the current ticker
        ticker = yf.Ticker(ticker_symbol)
        
        # Fetch the income statement and balance sheet
        income_statement = ticker.financials
        balance_sheet = ticker.balance_sheet
        
        # Define the years
        years = [2023, 2022, 2021, 2020]
        
        # Initialize lists to store the DuPont ratios for each year
        net_income_to_ebt_list = []
        ebt_to_ebit_list = []
        ebit_to_revenues_list = []
        revenues_to_assets_list = []
        assets_to_equity_list = []

        # Perform DuPont Analysis for each year
        for year in years:
            # Get the relevant financial data for the year
            net_income = income_statement.loc['Net Income', str(year)]
            interest_expense = income_statement.loc['Interest Expense', str(year)]
            ebit = income_statement.loc['EBIT', str(year)]
            total_revenue = income_statement.loc['Total Revenue', str(year)]
            total_assets = balance_sheet.loc['Total Assets', str(year)]
            stockholders_equity = balance_sheet.loc['Stockholders Equity', str(year)]

            # Calculate DuPont ratios
            net_income_to_ebt = net_income / (ebit - interest_expense)
            ebt_to_ebit = (ebit - interest_expense) / ebit
            ebit_to_revenues = ebit / total_revenue
            revenues_to_assets = total_revenue / total_assets
            assets_to_equity = total_assets / stockholders_equity

            # Append the calculated ratios to the respective lists
            net_income_to_ebt_list.append(net_income_to_ebt)
            ebt_to_ebit_list.append(ebt_to_ebit)
            ebit_to_revenues_list.append(ebit_to_revenues)
            revenues_to_assets_list.append(revenues_to_assets)
            assets_to_equity_list.append(assets_to_equity)

        # Assign the lists of ratios to the corresponding rows in the DataFrame
        dupont_breakdown_df.loc[ticker_symbol] = [net_income_to_ebt_list, ebt_to_ebit_list, ebit_to_revenues_list, revenues_to_assets_list, assets_to_equity_list]

    except Exception as e:
        print(f"Error fetching data for {ticker_symbol}: {e}")

# Transpose the DataFrame so that the years become columns
dupont_breakdown_df = dupont_breakdown_df.transpose()

# Display the transposed DuPont Analysis breakdown DataFrame
print(dupont_breakdown_df)

# Re-initialize the DataFrame with the correct columns
metrics = ["Net_Income_to_EBT", "EBT_to_EBIT", "EBIT_to_Revenues", "Revenues_to_Assets", "Assets_to_Equity"]
years = [2023, 2022, 2021, 2020]
ticker_symbols = ["INTC", "GOOG", "AAPL", "MSFT", "KO", "TSLA", "NVDA", "AMZN", "AVGO", "V", "WMT", "UNH", "HD", "OXY", "JNJ", "COST", "CVX", "AMD", "MCD", "CSCO", "ABT", "CAT", "TXN", "PFE", "PM", "UPS", "T", "VZ", "DE", "BA", "LMT", "SBUX"]

# Create a list of all column names needed
column_names = [f"{ticker} {year}" for ticker in ticker_symbols for year in years]

# Initialize structured_df with the correct column names
structured_df = pd.DataFrame(index=metrics, columns=column_names)

# Ensure each metric's value is a scalar by explicitly converting it if necessary
for ticker_symbol in ticker_symbols:
    for year in years:
        for metric in metrics:
            year_index = years.index(year)
            try:
                metric_series = dupont_breakdown_df.loc[metric, ticker_symbol][year_index]
                
                # Check if the series is empty or if it's a single value
                if isinstance(metric_series, pd.Series) and not metric_series.empty:
                    metric_value = metric_series.iloc[0]
                else:
                    metric_value = np.nan  # Assign NaN if the series is empty or not a series
                
                column_name = f"{ticker_symbol} {year}"
                structured_df.loc[metric, column_name] = metric_value
                
            except IndexError:
                # Handle the case where the series is empty or out of bounds
                print(f"Data not found for {ticker_symbol}, {metric}, {year}.")
                structured_df.loc[metric, column_name] = np.nan  # Assign NaN if data is not available

structured_df.head()  # Display the first few rows to verify

structured_df

# Calculate ROE and add it to structured_df
roe_row = []

# Loop through each column (ticker and year combination)
for column_name in structured_df.columns:
    product = 1  # Start with a product of 1
    for metric in metrics:
        value = structured_df.loc[metric, column_name]
        product *= value if not pd.isnull(value) else 0  # Multiply by value if not NaN, else multiply by 1
    
    roe_row.append(product)

# Add the ROE row to structured_df
structured_df.loc['ROE'] = roe_row

structured_df.head()  # This will now include the ROE at the bottom

# Excel file path - this names the file "Hackleman Resume Project.xlsx"
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'

# Check if the directory exists, if not, create it
directory = os.path.dirname(excel_path)
if not os.path.exists(directory):
    os.makedirs(directory)

# Create a new Excel file with an initial sheet
wb = Workbook()
ws = wb.active
ws.title = "ROE Calc"  # This names the first sheet within the workbook

# Save the workbook with the specified file name
wb.save(filename=excel_path)

# Write the DataFrame to the Excel file
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    structured_df.to_excel(writer, sheet_name='ROE Calc')

print("DuPont Analysis data for Portfolio has been saved to Excel.")

# Assuming the data provided is for ROE values
structured_df = pd.DataFrame(structured_df, index=["Net_Income_to_EBT", "EBT_to_EBIT", "EBIT_to_Revenues", "Revenues_to_Assets", "Assets_to_Equity", "ROE"])

# Isolating the ROE row
roe_values = structured_df.loc["ROE"]

# Creating a Series to hold the average ROE for each ticker
average_roe_per_ticker = pd.Series(dtype=float)

# Extracting ticker names and years, then calculating average ROE
for column in roe_values.index:
    ticker = column.split()[0]  # Splitting at space and taking the first part for ticker name
    if ticker in average_roe_per_ticker:
        average_roe_per_ticker[ticker] += roe_values[column]
    else:
        average_roe_per_ticker[ticker] = roe_values[column]

# Dividing by the number of years to get the average
average_roe_per_ticker = average_roe_per_ticker / 4

average_roe_per_ticker

# Initialize an empty DataFrame to store the dividend payout ratios
dividend_payout_ratio_df = pd.DataFrame(index=ticker_symbols, columns=["Average Dividend Payout Ratio", "Retention Ratio", "Flag for Review"])

for ticker in ticker_symbols:
    try:
        company = yf.Ticker(ticker)
        cash_flow = company.cashflow
        income_statement = company.financials

        dividends_paid = cash_flow.loc['Common Stock Dividend Paid'][:4].fillna(0)  # Last 4 years, fill NaN with 0
        net_income = income_statement.loc['Net Income'][:4].fillna(0)  # Last 4 years, fill NaN with 0

        # Use absolute value for dividends_paid to handle negative values typically reported in cash flows
        ratio = abs(dividends_paid / net_income)

        # Calculate the mean ratio, setting NaN values to 0 if net_income is 0 (to handle division by zero cases)
        average_ratio = ratio.mean()

        # Store the calculated average dividend payout ratio
        dividend_payout_ratio_df.at[ticker, 'Average Dividend Payout Ratio'] = average_ratio if not pd.isna(average_ratio) else 0

    except Exception as e:
        print(f"Error processing {ticker}: {e}")
        dividend_payout_ratio_df.at[ticker, 'Average Dividend Payout Ratio'] = 0

# Calculate the retention ratio by subtracting the average dividend payout ratio from 1
dividend_payout_ratio_df['Retention Ratio'] = 1 - dividend_payout_ratio_df['Average Dividend Payout Ratio']

# Ensure the retention ratio is not negative and not greater than 1
dividend_payout_ratio_df['Retention Ratio'] = dividend_payout_ratio_df['Retention Ratio'].clip(lower=0, upper=1)

# Add a column to flag any instances requiring further review
dividend_payout_ratio_df['Flag for Review'] = np.where(dividend_payout_ratio_df['Average Dividend Payout Ratio'] > 1, 'X', '')

print(dividend_payout_ratio_df)

# Convert the Series to a DataFrame
average_roe_df = average_roe_per_ticker.to_frame(name='Average ROE')

# Now merge with dividend_payout_ratio_df
merged_df = average_roe_df.merge(dividend_payout_ratio_df, left_index=True, right_index=True)

# Calculate growth for each ticker by multiplying the average ROE by the Retention Ratio
merged_df['Growth'] = merged_df['Average ROE'] * merged_df['Retention Ratio']

# Extract just the Growth column into a new DataFrame
growth_per_ticker = merged_df[['Growth']]

print(growth_per_ticker)

#Rename merged_df to forecast_fundamentals for clarity and use.
forecast_fundamentals = merged_df

# Specify the path to your Excel workbook
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'

# Use ExcelWriter with the mode 'a' to append to an existing file
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    forecast_fundamentals.to_excel(writer, sheet_name='Forecast Fundamentals')

print("Forecast fundamentals data for Portfolio has been saved to Excel.")

#BETA CALCULATION
# Define tickers
tickers = ["INTC", "GOOG", "AAPL", "MSFT", "KO", "TSLA", "NVDA", "AMZN", "AVGO", "V", "WMT", "UNH", "HD", "OXY", "JNJ", "COST", "CVX", "AMD", "MCD", "CSCO", "ABT", "CAT", "TXN", "PFE", "PM", "UPS", "T", "VZ", "DE", "BA", "LMT", "SBUX", "SPY"]

# Calculate start and end dates for 10 years
end_date = datetime.now()
start_date = end_date - timedelta(days=10*365)

# Download historical stock prices
data = yf.download(tickers, start=start_date.strftime('%Y-%m-%d'), end=end_date.strftime('%Y-%m-%d'))['Adj Close']

# Calculate daily returns
returns = data.pct_change()

# Calculate SPY variance
spy_variance = returns['SPY'].var()

# Calculate beta for each stock
beta_values = {}
for ticker in tickers[:-1]:  # Exclude SPY from the tickers list for this operation
    covariance = returns[ticker].cov(returns['SPY'])
    beta = covariance / spy_variance
    beta_values[ticker] = beta

# Convert beta values to a DataFrame for easy handling
beta_df = pd.DataFrame.from_dict(beta_values, orient='index', columns=['Beta'])

beta_df

#CALCULATING Expected Market Return
# Download historical stock prices for SPY
spy_data = yf.download("SPY", start=start_date, end=end_date)['Adj Close']

# Calculate daily returns for SPY
spy_daily_returns = spy_data.pct_change()

# Calculate the average daily return
average_daily_return = spy_daily_returns.mean()

# Convert the average daily return to an annualized return
annualized_return = (1 + average_daily_return) ** 252 - 1

print(f"Annualized Return for SPY over the last 10 years: {annualized_return:.2%}")

rfr = 0.0447

#CALCULATING OUR REQUIRED RETURN FOR EACH STOCK
# Initialize an empty DataFrame to store the required returns
req_ret_df = pd.DataFrame(index=beta_df.index, columns=['Required Return'])

# Iterate through the beta_df to calculate required return for each stock
for ticker in beta_df.index:
    beta = beta_df.loc[ticker, 'Beta']
    req_return = rfr + (beta * (annualized_return - rfr))
    req_ret_df.loc[ticker, 'Required Return'] = req_return

print(req_ret_df)

# Merge beta and required return DataFrames
capm_df = beta_df.merge(req_ret_df, left_index=True, right_index=True)

# Add Risk-Free Rate and Annualized Market Return as constants
capm_df['Risk-Free Rate'] = rfr
capm_df['Annualized Market Return'] = annualized_return

#Organize to clean up excel file
capm_df = capm_df[['Required Return', 'Risk-Free Rate', 'Beta', 'Annualized Market Return']]

# Export CAPM DataFrame to Excel
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    capm_df.to_excel(writer, sheet_name='CAPM Analysis')

print("CAPM analysis data for Portfolio has been saved to Excel.")

#PROJECT 2

def forecast_5_year_fcfe(ticker_symbol):
    company = yf.Ticker(ticker_symbol)
    income_statement = company.financials
    balance_sheet = company.balance_sheet
    cash_flow_statement = company.cashflow
    
    market_cap = company.info['marketCap']
    shares_outstanding = company.info['sharesOutstanding']

    revenues = income_statement.loc['Total Revenue'].iloc[:4][::-1]
    ebit = income_statement.loc['EBIT'].iloc[:4][::-1]
    tax_rate_for_calcs = income_statement.loc['Tax Rate For Calcs'].iloc[0]

    # Handling 'Purchase of PPE' and 'Capital Expenditure'
    if 'Purchase of PPE' in cash_flow_statement.index:
        purchase_of_ppe_series = abs(cash_flow_statement.loc['Purchase of PPE'].iloc[:4])
    elif 'Capital Expenditure' in cash_flow_statement.index:
        purchase_of_ppe_series = abs(cash_flow_statement.loc['Capital Expenditure'].iloc[:4])
    else:
        purchase_of_ppe_series = pd.Series([0, 0, 0, 0])

    # Direct mean calculation, falling back to 0 if not available
    depreciation_and_amortization = abs(cash_flow_statement.loc['Depreciation And Amortization'].iloc[:4].mean() if 'Depreciation And Amortization' in cash_flow_statement.index else 0)

    total_debt = balance_sheet.get('Total Debt', pd.Series([0])).iloc[0]
    current_assets = balance_sheet.get('Current Assets', pd.Series([0])).iloc[0]
    cash_and_equivalents = balance_sheet.get('Cash And Cash Equivalents', pd.Series([0])).iloc[0] + balance_sheet.get('Cash Cash Equivalents And Short Term Investments', pd.Series([0])).iloc[0]
    current_liabilities = balance_sheet.get('Current Liabilities', pd.Series([0])).iloc[0]

    adjusted_working_capital = (current_assets - cash_and_equivalents) - current_liabilities if current_liabilities > 0 else 0

    cagr = (revenues.iloc[-1] / revenues.iloc[0]) ** (1 / (len(revenues) - 1)) - 1
    average_ebit_margin = ebit.mean() / revenues.mean()
    debt_ratio = total_debt / (total_debt + market_cap) if total_debt + market_cap != 0 else 0.2

    # If the calculated debt_ratio is 0, set it to 0.2
    debt_ratio = 0.2 if debt_ratio == 0 else debt_ratio

    # Calculating yearly ratios
    yearly_capex_ratios = [ppe / rev if rev != 0 else 0 for ppe, rev in zip(purchase_of_ppe_series, revenues)]
    yearly_dep_ratios = [depreciation_and_amortization / rev if rev != 0 else 0 for rev in revenues]
    yearly_wcinv_ratios = [adjusted_working_capital / rev if rev != 0 else 0 for rev in revenues]

    average_capex = np.mean(yearly_capex_ratios)
    average_dep = np.mean(yearly_dep_ratios)
    average_wcinv = np.mean(yearly_wcinv_ratios)

    forecasted_revenues = [revenues.iloc[-1] * (1 + cagr) ** i for i in range(1, 6)]
    forecasted_ebit = [rev * average_ebit_margin for rev in forecasted_revenues]
    forecasted_capex = [rev * average_capex for rev in forecasted_revenues]
    forecasted_dep = [rev * average_dep for rev in forecasted_revenues]
    forecasted_wcinv = [rev * average_wcinv for rev in forecasted_revenues]
    
    forecasted_fcfe = [forecasted_ebit[i] * (1 - tax_rate_for_calcs) +
                       ((forecasted_capex[i] - forecasted_dep[i]) * (1 - debt_ratio)) -
                       (forecasted_wcinv[i] * (1 - debt_ratio)) for i in range(5)]
    
    fcfe_per_share_forecast = [fcfe / shares_outstanding for fcfe in forecasted_fcfe]

    forecast_df = pd.DataFrame({
        'Year': range(1, 6),
        'FCFE per Share': fcfe_per_share_forecast
    }).set_index('Year')

    detailed_forecast_df = pd.DataFrame({
        'Forecasted Revenue': forecasted_revenues,
        'Forecasted EBIT': forecasted_ebit,
        'Forecasted Capex': forecasted_capex,
        'Forecasted Dep': forecasted_dep,
        'Forecasted Working Capital Investment': forecasted_wcinv,
        'Year': range(1, 6)
    }).set_index('Year')
    
    # Include additional variables in the DataFrame
    detailed_forecast_df['Tax Rate for Calcs'] = tax_rate_for_calcs
    detailed_forecast_df['Growth Rate'] = cagr
    detailed_forecast_df['Required Return'] = req_return  # Assume req_return is defined elsewhere
    detailed_forecast_df['Debt Ratio'] = debt_ratio
    detailed_forecast_df['Shares'] = shares_outstanding
    
    # Return both the FCFE per share forecast and the detailed forecast DataFrame
    return forecast_df, detailed_forecast_df

# Define the list of tickers
tickers = ["INTC", "GOOG", "AAPL", "MSFT", "KO", "TSLA", "NVDA", "AMZN", "AVGO", "V", "WMT", "UNH", "HD", "OXY", "JNJ", "COST", "CVX", "AMD", "MCD", "CSCO", "ABT", "CAT", "TXN", "PFE", "PM", "UPS", "T", "VZ", "DE", "BA", "LMT", "SBUX"]

forecast_dfs = {}

excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'

# Iterate through the tickers and call the forecast function for each
for ticker in tickers:
    # Unpack the two DataFrames returned by the function
    fcfe_forecast, detailed_forecast = forecast_5_year_fcfe(ticker)
    
    forecast_dfs[ticker] = fcfe_forecast
    
    if fcfe_forecast['FCFE per Share'].min() < 0:
        # Use ExcelWriter with mode 'a' to append to an existing workbook
        # Dynamically name the sheet based on the ticker
        sheet_name = f"FCFE Review - {ticker}"
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            detailed_forecast.to_excel(writer, sheet_name=sheet_name)
    
    print(f"Forecast for {ticker}:")
    print(fcfe_forecast)
    print("\n")

print("Detailed FCFE reviews for tickers with negative forecast have been saved to Excel.")

# Optional: Combine into a multi-index DataFrame for a consolidated view
# Create an empty list to collect the DataFrames
df_list = []

# Iterate through the forecast DataFrames and add a ticker level to the index
for ticker, df in forecast_dfs.items():
    df['Ticker'] = ticker
    df.set_index('Ticker', append=True, inplace=True)
    df_list.append(df)

# Concatenate all DataFrames along the rows
consolidated_forecast_df = pd.concat(df_list)

# Reorder the index levels so Ticker comes first
consolidated_forecast_df = consolidated_forecast_df.reorder_levels(['Ticker', 'Year'])

print("Consolidated 5-Year FCFE Forecast for Multiple Tickers:")
print(consolidated_forecast_df)

# Define the list of tickers
tickers = ["INTC", "GOOG", "AAPL", "MSFT", "KO", "TSLA", "NVDA", "AMZN", "AVGO", "V", "WMT", "UNH", "HD", "OXY", "JNJ", "COST", "CVX", "AMD", "MCD", "CSCO", "ABT", "CAT", "TXN", "PFE", "PM", "UPS", "T", "VZ", "DE", "BA", "LMT", "SBUX"]

# Step 1: Initialize an empty DataFrame to store Terminal Values
terminal_values = pd.DataFrame(columns=['Ticker', 'Year', 'FCFE per Share'])

# Iterate through each ticker to calculate its Terminal Value
for ticker in tickers[:-1]:  # Excluding SPY from the ticker list for this operation
    year_5_fcfe = consolidated_forecast_df.loc[(ticker, 5), 'FCFE per Share']
    growth = growth_per_ticker.loc[ticker, 'Growth']
    req_ret = req_ret_df.loc[ticker, 'Required Return']
    
    # Adjust req_ret if it's less than growth
    adjusted_req_ret = req_ret if req_ret > growth else growth + 0.05

    # Calculate Terminal Value with the adjusted required return
    tv = (year_5_fcfe * (1 + growth)) / (adjusted_req_ret - growth)
    
    # Create a temporary DataFrame for the terminal value of the current ticker
    temp_df = pd.DataFrame({'Ticker': [ticker], 'Year': ['TV'], 'FCFE per Share': [tv]})
    
    # Concatenate this temporary DataFrame with the 'terminal_values' DataFrame
    terminal_values = pd.concat([terminal_values, temp_df], ignore_index=True)

# Now that all terminal values are in 'terminal_values', set its index for merging
terminal_values.set_index(['Ticker', 'Year'], inplace=True)

# Concatenate the terminal values to the forecast dataframe
consolidated_with_tv_df = pd.concat([consolidated_forecast_df, terminal_values])

# Assuming 'consolidated_with_tv_df' and 'req_ret_df' are correctly calculated
# Initialize an empty DataFrame to store the present value of FCFE for each ticker
pv_of_fcfe = pd.DataFrame(columns=['Ticker', 'PV of FCFE'])

# Calculate present value of FCFE for each stock directly, ensuring index integrity
pv_of_fcfe_dict = {}
for ticker in tickers[:-1]:  # Excluding SPY
    if ticker in consolidated_with_tv_df.index.get_level_values(0):
        req_return = req_ret_df.at[ticker, 'Required Return']
        total_pv = 0
        for year in range(1, 6):
            if (ticker, year) in consolidated_with_tv_df.index:
                fcfe = consolidated_with_tv_df.at[(ticker, year), 'FCFE per Share']
                total_pv += fcfe / ((1 + req_return) ** year)
        if (ticker, 'TV') in consolidated_with_tv_df.index:
            tv = consolidated_with_tv_df.at[(ticker, 'TV'), 'FCFE per Share']
            total_pv += tv / ((1 + req_return) ** 5)
        pv_of_fcfe_dict[ticker] = total_pv

pv_of_fcfe = pd.DataFrame(list(pv_of_fcfe_dict.items()), columns=['Ticker', 'PV of FCFE']).set_index('Ticker')
print(pv_of_fcfe)

# Export FCFE DataFrame to Excel
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    pv_of_fcfe.to_excel(writer, sheet_name='FCFE Analysis')

print("FCFE analysis data for Portfolio has been saved to Excel.")

#Scenario Analysis

# Initialize a DataFrame to hold scenario analysis results
# Ensure pv_of_fcfe['PV of FCFE'] is numeric and contains no NaN values
pv_of_fcfe['PV of FCFE'] = pd.to_numeric(pv_of_fcfe['PV of FCFE'], errors='coerce').fillna(0)

# Define the percentage variation for the scenarios
percentage_variation = 0.1  # 10% for both low and high cases

# Calculate Low and High Cases
pv_of_fcfe['Low Case'] = pv_of_fcfe['PV of FCFE'] * (1 - percentage_variation)
pv_of_fcfe['High Case'] = pv_of_fcfe['PV of FCFE'] * (1 + percentage_variation)

# Calculate the Range
pv_of_fcfe['Range'] = pv_of_fcfe['High Case'] - pv_of_fcfe['Low Case']

# Check the DataFrame to ensure the calculations are correct
print(pv_of_fcfe)

# Export the updated DataFrame with scenario analysis to Excel
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    pv_of_fcfe.to_excel(writer, sheet_name='Scenario Analysis')

print("Scenario analysis data for Portfolio has been updated and saved to Excel.")

#PROJECT 3
# Define tickers and target return
tickers = ["INTC", "GOOG", "AAPL", "MSFT", "KO", "TSLA", "NVDA", "AMZN", "AVGO", "V", "WMT", "UNH", "HD", "OXY", "JNJ", "COST", "CVX", "AMD", "CRM", "MCD", "CSCO", "ABT", "CAT", "TXN", "PFE", "NKE", "PM", "UPS", "T", "VZ", "DE", "BA", "LMT", "SBUX"]
tickers.append("SPY")  # Adding benchmark
target_ret = 0.20

# Calculate date 10 years ago from today
start_date = (datetime.now() - timedelta(days=365.25*10)).strftime('%Y-%m-%d')
end_date = datetime.now().strftime('%Y-%m-%d')

# Fetch historical data from Yahoo Finance
data = yf.download(tickers, start=start_date, end=end_date)['Adj Close']

print(data.head())

# Calculate daily returns
daily_returns = data.pct_change().dropna()

# Separate SPY for benchmark comparison
spy_daily_returns = daily_returns["SPY"]
daily_returns = daily_returns.drop(columns=["SPY"])

def objective(weights):
    port_ret = np.dot(daily_returns.mean(), weights) * 252  # Annualized
    port_vol = np.sqrt(np.dot(weights.T, np.dot(daily_returns.cov() * 252, weights)))
    sharpe_ratio = port_ret / port_vol
    return -sharpe_ratio

# Constraints: weights sum to 1, and portfolio return equals target return
constraints = (
    {'type': 'eq', 'fun': lambda weights: np.sum(weights) - 1},
    {'type': 'eq', 'fun': lambda weights: np.dot(daily_returns.mean(), weights) * 252 - target_ret}
)

# Bounds for each weight
bounds = tuple((0.01, 1) for _ in range(len(tickers) - 1))  # -1 because we removed SPY

# Initial guess
initial_guess = [1. / (len(tickers) - 1)] * (len(tickers) - 1)

# Optimization
result = minimize(objective, initial_guess, method='SLSQP', bounds=bounds, constraints=constraints)

# Resulting optimal weights
optimal_weights = result.x

# Calculate portfolio returns
portfolio_returns = daily_returns.dot(optimal_weights)

# Plotting portfolio returns vs. SPY
plt.figure(figsize=(14, 7))
plt.plot((1 + portfolio_returns).cumprod(), label='Optimal Portfolio')
plt.plot((1 + spy_daily_returns).cumprod(), label='SPY Benchmark')
plt.title('Optimal Portfolio vs. SPY')
plt.legend()
plt.show()

# Create a DataFrame for weights
weights_df = pd.DataFrame(optimal_weights, index=daily_returns.columns, columns=['Optimal Weights'])

# Calculate the daily returns of the portfolio
portfolio_daily_returns = daily_returns.dot(optimal_weights)

# Create a new DataFrame for Daily Portfolio Returns
portfolio_returns_df = pd.DataFrame(portfolio_daily_returns, columns=['Daily Portfolio Returns'])

# Show the first few rows of the new DataFrame
print(portfolio_returns_df.head())

# Specify the path to your Excel workbook
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'

# Use ExcelWriter with the mode 'a' to append to an existing file
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    data.to_excel(writer, sheet_name='Stock Prices Last 10 Years')

print("Stock prices for the last 10 years have been added to 'Hackleman Resume Project.xlsx'.")

# Combine portfolio returns and SPY daily returns into a single DataFrame
combined_returns_df = pd.DataFrame({
    'Portfolio Daily Returns': portfolio_daily_returns,
    'SPY Daily Returns': spy_daily_returns
})

# Exporting combined returns to Excel
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    combined_returns_df.to_excel(writer, sheet_name='Daily Returns & Benchmark')

print("Daily portfolio returns and SPY benchmark returns have been added to 'Hackleman Resume Project.xlsx'.")

plt.figure(figsize=(14, 7))
plt.plot((1 + portfolio_daily_returns).cumprod(), label='Optimal Portfolio')
plt.plot((1 + spy_daily_returns).cumprod(), label='SPY Benchmark')
plt.title('Optimal Portfolio vs. SPY')
plt.legend()
plt.savefig('optimal_vs_spy.png')
plt.close()

# Load the existing workbook and select a sheet
wb = load_workbook(excel_path)
ws = wb.create_sheet('Backtesting Graph')

# Insert the image into the sheet
img = Image('optimal_vs_spy.png')
ws.add_image(img, 'A1')  # Adjust 'A1' as needed based on where you want the image

# Save the workbook
wb.save(excel_path)
wb.close()

print("Backtesting graph has been added to 'Hackleman Resume Project.xlsx'.")

# Assuming weights_df contains the optimized weights as decimals
weights_df['Optimal Weights'] = weights_df['Optimal Weights'] * 100

# Optionally, you can round the percentages for cleaner presentation
weights_df['Optimal Weights'] = weights_df['Optimal Weights'].round(2)

# Now, weights_df contains the weights in percentage terms
print(weights_df.head())

# Exporting to Excel as before
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    weights_df.to_excel(writer, sheet_name='Optimized Weights in %')

print("Optimized portfolio weights in percentage terms have been added to 'Hackleman Resume Project.xlsx'.")

#PROJECT 4

# Calculate average returns (mean) for portfolio and SPY
mean_port_ret = portfolio_returns_df['Daily Portfolio Returns'].mean()
mean_spy_ret = spy_daily_returns.mean()

# Calculate standard deviation for portfolio and SPY
stdev_port = portfolio_returns_df['Daily Portfolio Returns'].std()
stdev_spy = spy_daily_returns.std()

# Portfolio value
portfolio_value = 10000000  # $10,000,000

# Confidence levels for VaR
z_5 = -1.645  # 5% VaR
z_1 = -2.326  # 1% VaR

# Calculate VaR in percentage terms
VaR_5_port = (mean_port_ret + z_5 * stdev_port)*-1
VaR_1_port = (mean_port_ret + z_1 * stdev_port)*-1

# Calculate VaR in dollar terms
VaR_5_port_dollar = VaR_5_port * portfolio_value
VaR_1_port_dollar = VaR_1_port * portfolio_value

# Repeat calculations for SPY
VaR_5_spy = (mean_spy_ret + z_5 * stdev_spy) * -1
VaR_1_spy = (mean_spy_ret + z_1 * stdev_spy) * -1
VaR_5_spy_dollar = VaR_5_spy * portfolio_value 
VaR_1_spy_dollar = VaR_1_spy * portfolio_value 

# Sort returns in ascending order
sorted_port_rets = portfolio_returns_df['Daily Portfolio Returns'].sort_values()
sorted_spy_rets = spy_daily_returns.sort_values()

# Calculate historical VaR
hist_VaR_5_port = sorted_port_rets.quantile(0.05)*-1
hist_VaR_1_port = sorted_port_rets.quantile(0.01)*-1

# Calculate historical VaR in dollar terms
hist_VaR_5_port_dollar = hist_VaR_5_port * portfolio_value * -1
hist_VaR_1_port_dollar = hist_VaR_1_port * portfolio_value * -1

# Repeat calculations for SPY
hist_VaR_5_spy = sorted_spy_rets.quantile(0.05) * -1
hist_VaR_1_spy = sorted_spy_rets.quantile(0.01) * -1
hist_VaR_5_spy_dollar = hist_VaR_5_spy * portfolio_value 
hist_VaR_1_spy_dollar = hist_VaR_1_spy * portfolio_value

var_results = pd.DataFrame({
    'Metric': ['5% VaR %', '1% VaR %', '5% VaR $', '1% VaR $'],
    'Portfolio': [VaR_5_port, VaR_1_port, VaR_5_port_dollar, VaR_1_port_dollar],
    'Portfolio (Hist)': [hist_VaR_5_port, hist_VaR_1_port, hist_VaR_5_port_dollar, hist_VaR_1_port_dollar],
    'SPY': [VaR_5_spy, VaR_1_spy, VaR_5_spy_dollar, VaR_1_spy_dollar],
    'SPY (Hist)': [hist_VaR_5_spy, hist_VaR_1_spy, hist_VaR_5_spy_dollar, hist_VaR_1_spy_dollar]
})

# Specify the path to your Excel workbook
excel_path = r'C:\Users\mhack\OneDrive\Documents\Investment Analysis\Hackleman Resume Project.xlsx'

# Use ExcelWriter with the mode 'a' to append to an existing file
with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
    var_results.to_excel(writer, sheet_name='VaR Calcs', index=False)

print("VaR calculations have been added to 'Hackleman Resume Project.xlsx'.")

#Project 5
# Sample placeholders for annualized returns and standard deviation
annualized_port_return = portfolio_returns.mean() * 252
annualized_benchmark_return = spy_daily_returns.mean() * 252
annualized_port_std = portfolio_returns.std() * np.sqrt(252)

# Assuming RFR (Risk-Free Rate) is known
RFR = 0.04  # Example value, adjust based on current rates

# Calculate Portfolio Beta
covariance = portfolio_returns.cov(spy_daily_returns) * 252
benchmark_variance = spy_daily_returns.var() * 252
portfolio_beta = covariance / benchmark_variance

# Alpha Calculation
alpha = annualized_port_return - (portfolio_beta * annualized_benchmark_return)

# Sharpe Ratio Calculation
sharpe_ratio = (annualized_port_return - RFR) / annualized_port_std

# Information Ratio Calculation
active_return = annualized_port_return - annualized_benchmark_return
active_risk = portfolio_returns.sub(spy_daily_returns, axis=0).std() * np.sqrt(252)
information_ratio = active_return / active_risk

# Calculate cumulative returns, then convert to cumulative portfolio value assuming a starting value (e.g., 100)
portfolio_value_over_time = (1 + portfolio_daily_returns).cumprod() * 100  # Starting with a base value of 100

# Ensure the index is a datetime type for proper plotting
portfolio_value_over_time.index = pd.to_datetime(portfolio_value_over_time.index)

# Calculate moving averages of the portfolio value
ma_20 = portfolio_value_over_time.rolling(window=20).mean()
ma_50 = portfolio_value_over_time.rolling(window=50).mean()
ma_200 = portfolio_value_over_time.rolling(window=200).mean()

import matplotlib.pyplot as plt

plt.figure(figsize=(14, 7))
plt.plot(portfolio_value_over_time.index, portfolio_value_over_time, label='Portfolio Value', color='black')
plt.plot(ma_20.index, ma_20, label='20-Day MA', color='red')
plt.plot(ma_50.index, ma_50, label='50-Day MA', color='blue')
plt.plot(ma_200.index, ma_200, label='200-Day MA', color='green')
plt.title('Portfolio Value and Moving Averages')
plt.legend()
plt.show()

#Save as an image
plt.figure(figsize=(14, 7))
plt.plot(portfolio_value_over_time.index, portfolio_value_over_time, label='Portfolio Value', color='black')
plt.plot(ma_20.index, ma_20, label='20-Day MA', color='red')
plt.plot(ma_50.index, ma_50, label='50-Day MA', color='blue')
plt.plot(ma_200.index, ma_200, label='200-Day MA', color='green')
plt.title('Portfolio Value and Moving Averages')
plt.legend()

plt.savefig('moving_averages_plot.png')

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Load the workbook
wb = load_workbook(excel_path)

# Check if the "Chart" sheet exists; if not, create it
if "Chart" not in wb.sheetnames:
    ws = wb.create_sheet("MVA Chart")
else:
    ws = wb["MVA Chart"]

# Load and insert the image
img = Image('moving_averages_plot.png')
ws.add_image(img, 'A1')  # Adjust the cell as needed

# Save the workbook
wb.save(excel_path)
wb.close()

print("The moving averages chart has been saved to the 'MVA Chart' sheet in 'Hackleman Resume Project.xlsx'.")


metrics_df = pd.DataFrame({
    'Metric': ['Alpha', 'Sharpe Ratio', 'Information Ratio'],
    'Value': [alpha, sharpe_ratio, information_ratio]
})

# Exporting to Excel
try:
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        metrics_df.to_excel(writer, sheet_name='Fund-Tech Analysis')
    print("Metrics have been successfully exported to 'Hackleman Resume Project.xlsx'.")
except PermissionError as e:
    print("Permission denied error: Make sure the file is not open in another program.", e)
except ValueError as e:
    print("Value error occurred:", e)





